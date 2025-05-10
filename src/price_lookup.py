# price_lookup.py
import requests
from bs4 import BeautifulSoup
import re
import time
from openpyxl import load_workbook

def fetch_ebay_price(card_name):
    """
    Search eBay sold listings for the card name and return average price of top 3 results.
    Returns None if no price found.
    """
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
    }

    query = requests.utils.quote(card_name + " trading card")
    url = f"https://www.ebay.com/sch/i.html?_nkw={query}&_sacat=0&LH_Sold=1&LH_Complete=1"

    try:
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
    except requests.RequestException as e:
        print(f"Error fetching eBay results for {card_name}: {e}")
        return None

    soup = BeautifulSoup(response.text, "html.parser")
    prices = []

    for item in soup.select(".s-item"):
        price_tag = item.select_one(".s-item__price")
        if price_tag:
            text = price_tag.get_text()
            match = re.search(r"\$([\d,.]+)", text)
            if match:
                price = float(match.group(1).replace(",", ""))
                prices.append(price)
        if len(prices) >= 3:
            break

    if prices:
        return round(sum(prices) / len(prices), 2)
    return None

def update_excel_with_prices(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    if "Last Sold Price (eBay)" not in header:
        ws.cell(row=1, column=len(header) + 1).value = "Last Sold Price (eBay)"
        price_col = len(header) + 1
    else:
        price_col = header.index("Last Sold Price (eBay)") + 1

    for row in ws.iter_rows(min_row=2):
        card_name = row[0].value
        if card_name:
            price = fetch_ebay_price(card_name)
            print(f"{card_name} → ${price if price else 'N/A'}")
            ws.cell(row=row[0].row, column=price_col).value = price
            time.sleep(2)  # be gentle with eBay

    wb.save(excel_path)
    return excel_path
